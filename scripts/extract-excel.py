#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
礼拝出席名簿_2026集計用.xlsx を読み取り、中間 JSON を生成する（読み取り専用）。
元 Excel は一切変更しない（openpyxl を read_only で開き保存処理を書かない）。

使い方:
    python3 scripts/extract-excel.py "<xlsx パス>" [出力先json]

出力（既定）: scripts/_artifacts/excel-extract.json
個人情報（氏名）を含むため、_artifacts/ は .gitignore 済み。使用後の削除を推奨。

後段の scripts/migrate-from-excel.ts がこの JSON を読み、Synaxis へ移行する。
"""
import sys, os, json, re, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です。 pip install openpyxl を実行してください。")

# シート名 -> (relationship_status, age_group)
ROSTER_MAP = {
    "会員": ("member", "adult"),
    "客員": ("regular_attendee", "adult"),
    "未信": ("seeker", "adult"),
    "子ども(小3以下)": ("member", "child"),  # 立場は要レビュー（暫定 member）
    "ビジター": ("guest", "adult"),
    "ビジター2": ("guest", "adult"),
}
MARK = "○"

def iso(v):
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    return None

def parse_ymd(s):
    if s is None:
        return None
    if isinstance(s, (datetime.datetime, datetime.date)):
        return iso(s)
    m = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", str(s))
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None

def extract_roster(ws, relationship, age_group):
    rows = list(ws.iter_rows(values_only=True))
    # 見出し行 = 日付(datetime)セルが最も多い行
    header_idx, header = -1, None
    best = 0
    for i, r in enumerate(rows[:12]):
        n = sum(1 for c in (r or []) if isinstance(c, (datetime.datetime, datetime.date)))
        if n > best:
            best, header_idx, header = n, i, r
    if header_idx < 0:
        return [], []
    # 列 -> 日付
    date_cols = {}
    for ci, c in enumerate(header):
        d = iso(c)
        if d:
            date_cols[ci] = d
    name_col = 2  # C列(0始まりで2)
    num_col = 1   # B列(#)
    people = []
    for r in rows[header_idx + 1:]:
        if r is None:
            continue
        num = r[num_col] if len(r) > num_col else None
        name = r[name_col] if len(r) > name_col else None
        if not isinstance(num, (int, float)):
            continue  # # が数値でなければ人物行ではない
        if not isinstance(name, str) or not name.strip():
            continue
        dates = {date_cols[ci] for ci in date_cols
                 if len(r) > ci and isinstance(r[ci], str) and r[ci].strip() == MARK}
        people.append({
            "name": name.strip(),
            "relationship": relationship,
            "ageGroup": age_group,
            "attendanceDates": sorted(dates),
        })
    return people, sorted(date_cols.values())

def extract_special(ws):
    rows = list(ws.iter_rows(values_only=True))
    # 見出し行: 'お名前' を含む行
    header_idx = -1
    for i, r in enumerate(rows[:8]):
        if r and any(isinstance(c, str) and "お名前" in c for c in r):
            header_idx = i
            break
    if header_idx < 0:
        return []
    header = rows[header_idx]
    name_cols = [ci for ci, c in enumerate(header) if isinstance(c, str) and "お名前" in c]
    events = []
    for nc in name_cols:
        mark_col, status_col = nc + 1, nc + 2
        # イベント日付: 見出しより上の行で nc 付近のセルから年月日を拾う
        date = None
        for i in range(0, header_idx):
            for ci in range(max(0, nc - 2), nc + 3):
                if rows[i] and len(rows[i]) > ci:
                    d = parse_ymd(rows[i][ci])
                    if d:
                        date = d
                        break
            if date:
                break
        month = int(date[5:7]) if date else 0
        name = "元旦礼拝" if month == 1 else "受難日礼拝" if month == 4 else f"特別礼拝 {date}"
        attendees = []
        for r in rows[header_idx + 1:]:
            if r is None or len(r) <= status_col:
                continue
            nm = r[nc]
            if not isinstance(nm, str) or not nm.strip():
                continue
            mark = r[mark_col] if len(r) > mark_col else None
            if not (isinstance(mark, str) and mark.strip() == MARK):
                continue
            status = r[status_col] if isinstance(r[status_col], str) else None
            attendees.append({"name": nm.strip(), "status": (status or "").strip()})
        if date:
            events.append({"eventName": name, "date": date, "kind": "special_worship",
                           "attendees": attendees})
    return events

def main():
    src = sys.argv[1] if len(sys.argv) > 1 else \
        os.path.expanduser("~/Documents/礼拝出席名簿_2026集計用-1.xlsx")
    out = sys.argv[2] if len(sys.argv) > 2 else \
        os.path.join(os.path.dirname(__file__), "_artifacts", "excel-extract.json")
    if not os.path.exists(src):
        sys.exit(f"ファイルが見つかりません: {src}")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    rosters = []
    weekly = set()
    for sheet, (rel, age) in ROSTER_MAP.items():
        if sheet not in wb.sheetnames:
            continue
        people, dates = extract_roster(wb[sheet], rel, age)
        weekly.update(dates)
        rosters.append({"sheet": sheet, "relationship": rel, "ageGroup": age,
                        "people": people})
    special = []
    for s in wb.sheetnames:
        if "元旦" in s or "受難" in s:
            special = extract_special(wb[s])
    wb.close()

    data = {
        "source": os.path.basename(src),
        "weeklyDates": sorted(weekly),
        "rosters": rosters,
        "special": special,
    }
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # サマリー（照合用）
    print(f"source: {data['source']}")
    print(f"weekly dates: {len(data['weeklyDates'])}")
    total_people = total_marks = 0
    for r in rosters:
        marks = sum(len(p['attendanceDates']) for p in r['people'])
        total_people += len(r['people'])
        total_marks += marks
        print(f"  {r['sheet']:14} people={len(r['people']):4} ○={marks}")
    print(f"  TOTAL people={total_people} ○={total_marks}")
    for e in special:
        print(f"  special {e['eventName']} {e['date']} attendees={len(e['attendees'])}")
    print(f"wrote: {out}")

if __name__ == "__main__":
    main()
